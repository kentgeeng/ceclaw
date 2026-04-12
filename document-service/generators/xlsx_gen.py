import httpx
import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import LLM_URL, LLM_MODEL

SEARXNG_URL = "http://127.0.0.1:2337/v1/search"

HEADER_FILL = PatternFill("solid", fgColor="0D0D2E")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Noto Serif CJK TC", size=11)
ALT_FILL = PatternFill("solid", fgColor="F0F0FA")
NORMAL_FONT = Font(name="Noto Serif CJK TC", size=10)
THIN = Side(style="thin", color="CCCCDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


async def search_web(query: str, num: int = 5) -> list:
    try:
        async with httpx.AsyncClient(timeout=15) as c:
            r = await c.post(SEARXNG_URL, json={"query": query})
            results = r.json().get("data", {}).get("web", [])[:num]
            return [{"title": x["title"], "description": x.get("description", "")} for x in results]
    except Exception:
        return []


async def llm_generate_table(topic: str, search_results: list) -> dict:
    context = "\n".join([f"- {r['title']}: {r['description'][:150]}" for r in search_results])
    prompt = f"""你是資深數據分析師，為「{topic}」生成一份專業的試算表資料。

網路最新參考資料：
{context}

規則：
- 標題精準描述內容
- 欄位名稱專業，包含數據單位
- 至少 8-12 行資料
- 資料要有真實感，包含具體數字、百分比、日期等
- 台灣本地化視角

只輸出 JSON：
{{
  "title": "表格標題",
  "subtitle": "資料說明（來源/時間）",
  "headers": ["欄位1", "欄位2", "欄位3"],
  "rows": [["值1", "值2", "值3"]],
  "summary": "數據摘要（一句話）"
}}"""

    async with httpx.AsyncClient(timeout=90) as c:
        r = await c.post(LLM_URL, json={
            "model": LLM_MODEL,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0,
            "max_tokens": 2000,
        })
    text = r.json()["choices"][0]["message"]["content"].strip()
    text = re.sub(r'<\|channel\|>.*?<channel\|>', '', text, flags=re.DOTALL)
    text = re.sub(r'<think>.*?</think>', '', text, flags=re.DOTALL)
    start = text.find("{")
    end = text.rfind("}") + 1
    return json.loads(text[start:end])


async def generate_xlsx(topic: str, output_path: str) -> str:
    search_results = await search_web(topic)
    data = await llm_generate_table(topic, search_results)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = data.get("title", "資料")[:31]

    # 標題列
    ws.merge_cells("A1:{}1".format(get_column_letter(len(data.get("headers", [1])))))
    title_cell = ws["A1"]
    title_cell.value = data.get("title", topic)
    title_cell.font = Font(name="Noto Serif CJK TC", size=14, bold=True, color="0D0D2E")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="E8E8F5")
    ws.row_dimensions[1].height = 28

    # 副標題
    col_count = len(data.get("headers", []))
    if col_count > 0:
        ws.merge_cells("A2:{}2".format(get_column_letter(col_count)))
        sub_cell = ws["A2"]
        sub_cell.value = data.get("subtitle", "CECLAW 資料服務")
        sub_cell.font = Font(name="Noto Serif CJK TC", size=9, color="888899")
        sub_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 18

    # 空行
    ws.row_dimensions[3].height = 8

    # 欄位標題
    headers = data.get("headers", [])
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(str(h)) * 2.2)
    ws.row_dimensions[4].height = 22

    # 資料列
    for row_idx, row in enumerate(data.get("rows", []), 5):
        is_alt = (row_idx % 2 == 0)
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if is_alt:
                cell.fill = ALT_FILL
        ws.row_dimensions[row_idx].height = 18

    # 摘要列
    if data.get("summary") and col_count > 0:
        sum_row = 5 + len(data.get("rows", []))
        ws.merge_cells("A{}:{}{}".format(sum_row, get_column_letter(col_count), sum_row))
        sum_cell = ws["A{}".format(sum_row)]
        sum_cell.value = "📊 " + data["summary"]
        sum_cell.font = Font(name="Noto Serif CJK TC", size=10, italic=True, color="555577")
        sum_cell.fill = PatternFill("solid", fgColor="EEEEFF")
        sum_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[sum_row].height = 20

    # 凍結標題
    ws.freeze_panes = "A5"

    wb.save(output_path)
    return output_path
